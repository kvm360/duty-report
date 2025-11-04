from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.db.models import Q
import pytz
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from .models import Shift, PTORequest, TeamMember
from .forms import CustomAuthenticationForm, ShiftForm, PTORequestForm, TeamMemberForm

def is_admin(user):
    return user.is_staff

def login_view(request):
    if request.method == 'POST':
        form = CustomAuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('dashboard')
    else:
        form = CustomAuthenticationForm()
    return render(request, 'scheduler/login.html', {'form': form})

def logout_view(request):
    logout(request)
    return redirect('login')

@login_required
def dashboard(request):
    now_utc = timezone.now()
    
    # Get current shifts
    current_shifts = Shift.objects.filter(
        start_time_utc__lte=now_utc,
        end_time_utc__gte=now_utc
    )
    
    # Get pending PTO requests count (for admin)
    pending_pto_count = 0
    if request.user.is_staff:
        pending_pto_count = PTORequest.objects.filter(status='Pending').count()
    
    # Get user's upcoming shifts
    user_shifts = Shift.objects.filter(member=request.user).filter(
        start_time_utc__gte=now_utc
    ).order_by('start_time_utc')[:5]
    
    context = {
        'current_shifts': current_shifts,
        'pending_pto_count': pending_pto_count,
        'user_shifts': user_shifts,
        'now_utc': now_utc,
    }
    
    if request.user.is_staff:
        # Admin dashboard - show all shifts for current week
        start_of_week = now_utc - timedelta(days=now_utc.weekday())
        end_of_week = start_of_week + timedelta(days=7)
        
        all_shifts = Shift.objects.filter(
            start_time_utc__gte=start_of_week,
            start_time_utc__lte=end_of_week
        ).order_by('start_time_utc')
        
        context['all_shifts'] = all_shifts
        context['week_start'] = start_of_week
        context['week_end'] = end_of_week
        
        return render(request, 'scheduler/admin_dashboard.html', context)
    else:
        # User dashboard
        return render(request, 'scheduler/user_dashboard.html', context)

@login_required
@user_passes_test(is_admin)
def add_shift(request):
    if request.method == 'POST':
        form = ShiftForm(request.POST)
        if form.is_valid():
            shift = form.save(commit=False)
            shift.created_by = request.user
            shift.save()
            return redirect('dashboard')
    else:
        form = ShiftForm()
    return render(request, 'scheduler/add_shift.html', {'form': form})

@login_required
@user_passes_test(is_admin)
def edit_shift(request, shift_id):
    shift = get_object_or_404(Shift, id=shift_id)
    if request.method == 'POST':
        form = ShiftForm(request.POST, instance=shift)
        if form.is_valid():
            form.save()
            return redirect('dashboard')
    else:
        form = ShiftForm(instance=shift)
    return render(request, 'scheduler/edit_shift.html', {'form': form, 'shift': shift})

@login_required
@user_passes_test(is_admin)
def delete_shift(request, shift_id):
    shift = get_object_or_404(Shift, id=shift_id)
    if request.method == 'POST':
        shift.delete()
        return redirect('dashboard')
    return render(request, 'scheduler/delete_shift.html', {'shift': shift})

@login_required
def my_schedule(request):
    # Get user's timezone
    try:
        user_tz = pytz.timezone(request.user.teammember.timezone)
    except (TeamMember.DoesNotExist, pytz.UnknownTimeZoneError):
        user_tz = pytz.UTC
    
    # Get shifts for current month
    now = timezone.now()
    start_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    if now.month == 12:
        end_of_month = now.replace(year=now.year + 1, month=1, day=1)
    else:
        end_of_month = now.replace(month=now.month + 1, day=1)
    
    shifts = Shift.objects.filter(
        member=request.user,
        start_time_utc__gte=start_of_month,
        start_time_utc__lte=end_of_month
    ).order_by('start_time_utc')
    
    # Convert to user's timezone for display
    for shift in shifts:
        shift.start_local = shift.start_time_utc.astimezone(user_tz)
        shift.end_local = shift.end_time_utc.astimezone(user_tz)
    
    context = {
        'shifts': shifts,
        'user_timezone': user_tz,
    }
    return render(request, 'scheduler/my_schedule.html', context)

@login_required
def all_members(request):
    members = User.objects.filter(is_staff=False).order_by('username')
    context = {'members': members}
    return render(request, 'scheduler/all_members.html', context)

@login_required
def member_schedule(request, username):
    member = get_object_or_404(User, username=username)
    
    # Get member's timezone
    try:
        member_tz = pytz.timezone(member.teammember.timezone)
    except (TeamMember.DoesNotExist, pytz.UnknownTimeZoneError):
        member_tz = pytz.UTC
    
    # Get shifts for current week
    now = timezone.now()
    start_of_week = now - timedelta(days=now.weekday())
    end_of_week = start_of_week + timedelta(days=7)
    
    shifts = Shift.objects.filter(
        member=member,
        start_time_utc__gte=start_of_week,
        start_time_utc__lte=end_of_week
    ).order_by('start_time_utc')
    
    # Convert to member's timezone for display
    for shift in shifts:
        shift.start_local = shift.start_time_utc.astimezone(member_tz)
        shift.end_local = shift.end_time_utc.astimezone(member_tz)
    
    context = {
        'member': member,
        'shifts': shifts,
        'member_timezone': member_tz,
    }
    return render(request, 'scheduler/member_schedule.html', context)

@login_required
def pto_requests(request):
    if request.method == 'POST' and not request.user.is_staff:
        form = PTORequestForm(request.POST)
        if form.is_valid():
            pto_request = form.save(commit=False)
            pto_request.user = request.user
            pto_request.save()
            return redirect('pto_requests')
    else:
        form = PTORequestForm()
    
    if request.user.is_staff:
        # Admin sees all PTO requests
        pto_requests_list = PTORequest.objects.all().order_by('-start_date')
    else:
        # User sees only their own PTO requests
        pto_requests_list = PTORequest.objects.filter(user=request.user).order_by('-start_date')
    
    context = {
        'form': form,
        'pto_requests': pto_requests_list,
    }
    return render(request, 'scheduler/pto_requests.html', context)

@login_required
@user_passes_test(is_admin)
def update_pto_status(request, pto_id):
    pto_request = get_object_or_404(PTORequest, id=pto_id)
    if request.method == 'POST':
        new_status = request.POST.get('status')
        if new_status in ['Approved', 'Rejected']:
            pto_request.status = new_status
            pto_request.save()
    return redirect('pto_requests')

@login_required
def export_schedule(request):
    # Get user's timezone
    try:
        user_tz = pytz.timezone(request.user.teammember.timezone)
    except (TeamMember.DoesNotExist, pytz.UnknownTimeZoneError):
        user_tz = pytz.UTC
    
    # Get shifts for current month
    now = timezone.now()
    start_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    if now.month == 12:
        end_of_month = now.replace(year=now.year + 1, month=1, day=1)
    else:
        end_of_month = now.replace(month=now.month + 1, day=1)
    
    shifts = Shift.objects.filter(
        member=request.user,
        start_time_utc__gte=start_of_month,
        start_time_utc__lte=end_of_month
    ).order_by('start_time_utc')
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "My Schedule"
    
    # Header row
    headers = ['Date', 'Start Time', 'End Time', 'Timezone', 'Title', 'Notes']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    for row, shift in enumerate(shifts, 2):
        start_local = shift.start_time_utc.astimezone(user_tz)
        end_local = shift.end_time_utc.astimezone(user_tz)
        
        ws.cell(row=row, column=1, value=start_local.date().isoformat())
        ws.cell(row=row, column=2, value=start_local.strftime('%I:%M %p'))
        ws.cell(row=row, column=3, value=end_local.strftime('%I:%M %p'))
        ws.cell(row=row, column=4, value=str(user_tz))
        ws.cell(row=row, column=5, value=shift.title)
        ws.cell(row=row, column=6, value=shift.notes or '')
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="schedule_{now.strftime("%Y_%m")}.xlsx"'
    wb.save(response)
    
    return response

@login_required
def profile_settings(request):
    try:
        team_member = request.user.teammember
    except TeamMember.DoesNotExist:
        team_member = TeamMember.objects.create(user=request.user, timezone='UTC')
    
    if request.method == 'POST':
        form = TeamMemberForm(request.POST, instance=team_member)
        if form.is_valid():
            form.save()
            return redirect('dashboard')
    else:
        form = TeamMemberForm(instance=team_member)
    
    return render(request, 'scheduler/profile_settings.html', {'form': form})
